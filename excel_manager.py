import os
from openpyxl import Workbook, load_workbook

FOLDER = "shortlisted"
LIMIT = 100

os.makedirs(FOLDER, exist_ok=True)


def save_candidate(candidate):

    file_no = 1

    while True:

        filename = os.path.join(FOLDER, f"shortlisted_{file_no}.xlsx")

        if not os.path.exists(filename):

            wb = Workbook()
            ws = wb.active

            ws.append([
                "Name",
                "Resume",
                "Overall Score",
                "Semantic Score",
                "Skill Score",
                "Recommendation"
            ])

            wb.save(filename)

        wb = load_workbook(filename)
        ws = wb.active

        # rows including header
        if ws.max_row <= LIMIT:
            ws.append([
                candidate["name"],
                candidate["resume"],
                candidate["overall_score"],
                candidate["semantic_score"],
                candidate["skill_score"],
                candidate["recommendation"]
            ])

            wb.save(filename)
            return filename

        file_no += 1